from xlutils.copy import copy
import xlwt
import urllib
from openpyxl import Workbook, load_workbook
#25 93
import zlib
import urllib, urllib2, cookielib
import json
try:
    from urllib.error import HTTPError
except:
    from urllib2 import HTTPError
from xlrd import open_workbook

a = (u'37 Robert-Koch-Stra\xdfe')
import ast
from lxml import html  
import csv,os,json
import requests
import re
print("yas")
import sys    
try:
    from urllib.request import Request, urlopen  # Python 3
except:
    from urllib2 import Request, urlopen

alpha_errors =[]
company_errors =[]
alphanum1_errors = []
alphanum2_errors = []
current_letter = ""
current_num1 = 0
current_num2 = 0
current_company = ""
total_comp = 63700
randomvar = 0
randomvar1 = 0
    
##username = 'jayj2019@gmail.com'
##password = 'clover1'
##

##login_data = urllib.urlencode({'session_key' : username, 'session_password' : password})
##opener.open('https://www.linkedin.com/uas/login', login_data)

import requests
from bs4 import BeautifulSoup

# Get login form

#cj = cookielib.CookieJar()
#opener = urllib2.build_opener(urllib2.HTTPCookieProcessor(cj))
URL = 'https://www.linkedin.com/uas/login'
session = requests.session()
login_response = session.get('https://www.linkedin.com/uas/login')
login = BeautifulSoup(login_response.text)

# Get hidden form inputs
inputs = login.find('form', {'name': 'login'}).findAll('input', {'type': ['hidden', 'submit']})

# Create POST data
post = {input.get('name'): input.get('value') for input in inputs}
post['session_key'] = 'jayj2019@gmail.com'
post['session_password'] = 'clover1'

# Post login
post_response = session.post('https://www.linkedin.com/uas/login-submit', data=post)

# Get home page
home_response = session.get('http://www.linkedin.com/nhome')
home = BeautifulSoup(home_response.text,"html5lib")

#cj = cookielib.CookieJar()
#opener = urllib2.build_opener(urllib2.HTTPCookieProcessor(cj))


#print(home)

#print resp.read()

#/div[@class="columns"]/ul[@class="column dual-column/li[@class="content"]/a
#['atlanta','baltimore','charlotte','chicago','cleveland','columbus-ohio','denver','houston','indianapolis','las-vegas','miami','milwaukee','oklahoma-city','orlando','phoenix','portland','provo','rochester'

#wb = Workbook()
#ws = wb.create_sheet()
##ws.title = 'Pi'
#wb.save(filename='book4.xls')
##
##
##book = xlwt.Workbook(encoding = "utf-8")
##sheet1 = book.add_sheet("Sheet 1")
##sheet1.write(0,0,"Company Name")
##sheet1.write(0,1,"Company Address")
##sheet1.write(0,2,"Company Website")
##sheet1.write(0,3,"Comapny Size")
##sheet1.write(0,4,"Company Industry Type")

def data():
 

            global randomvar1
            global total_comp
            times = 0
            global company_errors
            global current_company
            ecount = 0
            errcount = 0
            unierr = 0
            error_list = []
            xy = 0
            comp_count = 1000
                
            for count in range(comp_count,29999999):
       
                print("Count: "+str(count))
                try:
                    url = 'https://www.linkedin.com/company-beta/'+str(count)+'/'
                   
                    
 #                   req = Request(url, None, {'User-agent' : 'Mozilla/5.0 (Windows; U; Windows NT 5.1; de; rv:1.9.1.5) Gecko/20091102Firefox/3.5.5'})
 #                   urla = urllib2.urlopen(url)
 #                   resp4 = opener.open(req)
 #                   cont4 = resp4.read()
 #                   content = urla.read()
                    home_response = session.get(url)
                    home = BeautifulSoup(home_response.text,"lxml-xml")
                    print("url ,",url)
                    x = str(home)
                    doc4 = html.fromstring(x)
                    #print(content)
                    times+=1
                    #print(home)
 
                    
 #                  x = home.find(id="datalet-bpr-guid-1103794")
 #                   x = str(home)
 #                   tree = html.fromstring(x)
                    buyers = doc4.xpath('//code/text()')
                    o = 0
                    for a in buyers:
                        try:
 #                       unic = unicode(a)
                            y =a[3:-3]
     #                       jason = json.loads(y)
 #                           z = ast.literal_eval(y)
                            
                            #print(o)
                            o+=1
                            if "country:" in y or "industries:" in y or "companyPageUrl:" in y:
                                ind = y.find("city:")
                                ind1 = y.find("country:")
                                industries = y.find("industries:")
                                name = y.find("name:")
                                spec = y.find("specialities:")
                                line1_index = y.find("line1:")
                                line2_index = y.find("line2:")
                                site_index = y.find("companyPageUrl:")
                                postal_index = y.find("postalCode:")
                                type_index = y.find("type:")
                                #print(industries)
                                
                                city = ""
                                country = ""
                                indus = ""
                                nam= ""
                                speci = ""
                                line1 = ""
                                line2 = ""
                                site = ""
                                postal = ""
                                typec = ""
                                for i in y[ind:]:
                                    if i == ",":
                                        break
                                    city = city+i;
                                for i in y[ind1:]:
                                    if i == ",":
                                        break
                                    country = country+i;
                                for i in y[industries:]:
                                    if i == "]":
                                        break
                                    indus = indus+i
                                for i in y[name:]:
                                    if i == ",":
                                        break
                                    nam = nam+i;
                                for i in y[spec:]:
                                    if i == "]":
                                        break
                                    speci = speci+i;
                                if line1_index != -1:
                                    for i in y[line1_index:]:
                                        if i == "$":
                                            break
                                        line1 = line1+i;
                                if line2_index != -1:
                                    for i in y[line2_index:]:
                                        if i == ",":
                                            break
                                        line2 = line2+i;
                                if site_index != -1:
                                    for i in y[site_index:]:
                                        if i == ",":
                                            break
                                        site = site+i;
                                if postal_index != -1:
                                    for i in y[postal_index:]:
                                        if i == ",":
                                            break
                                        postal = postal+i
                                if type_index != -1:
                                    for i in y[type_index:]:
                                        if i == ",":
                                            break
                                        typec = typec+i
                                
                                                       #print "Name: " + name
##                        print "Address: " + address
##                        print "Website: " + website
##                        print "Company Size: " + sizeajdnajbffnfqnfwpnf
##                        print "Company Type: " + ctype
##                        print "Industry: " + indusfjfqfwfiffnifniw
                            
                                nam = nam[5:]
                                if indus[10] == ",":
                                    indus = "Industry not available"
                                else:
                                    indus = indus[12:]
                                if len(speci) == 14:
                                    speci = "Specializations not available"
                                else:
                                    speci = speci[14:]
                                if site == "":
                                    site = "Website not available"
                                else:
                                    site = site[15:]
                                if typec[5:25] == "com.linkedin.voyager":
                                    typec = "Company Type is not available"
                                else:
                                    typec = typec[5:]
                                address = line1[6:]+ line2[6:]+city[5:]+","+country[8:]+"-"+postal[11:]
                               
                                print("Name: "+nam)                                   
                                print("Address: "+address)
                                print("Industry: "+indus)
                                if speci == "":
                                    speci = "Specializations not available"
                                print("Specializations: "+speci)
                                print("Website: "+site)
                                print("Company Type: "+typec)
                                with open("companies.txt", "a") as text_file:
                                     text_file.write(nam + "^"+address+"^"+indus+"^"+speci+"^"+site+"^"+typec+"\n")
                                print()
                                print()
                           # print()
                        except SyntaxError:
                            print("syntax")
                            pass
 #                   times+=1      

            
                    
                    
            
# -> u'2 rooms \xb7 USD 0'
                    
                      
  
                   # print(buyers)
 #                   error =  home.xpath('//div[@class="alert error"]/p/strong/text()')
 
##                    if error:
##                        print("ERRRRRRRRRRRRRROR")
##                       
##                        #errcount+=1
##                    else:
##                        address = ""
##                        website = ""
##                        name = ""
##                        specialities = ""
##                        size = ""
##                        ctype = ""
##                        industry = ""
##                        
##                        print("hereeee")
##                        success = doc4.xpath('//code[@id="stream-about-section-embed-id-content"]/comment()')
##                        unic = unicode(success[0])
##                        string = (unic)[4:-3]
##                        compname = doc4.xpath('//meta[@property="og:title"]/@content')
##                        string = json.loads(string)
##                        name = str(compname[0])
##                        if 'headquarters' in string.keys():
##                            if 'street1' in string['headquarters'].keys() and len(str(string['headquarters']['street1']))!=0:
##                                address = address + str(string['headquarters']['street1'])+ ", "
##                            if 'street2' in string['headquarters'].keys() and len(str(string['headquarters']['street2']))!=0 :
##                                address = address + str(string['headquarters']['street2'])+ ", "
##                            if 'city' in string['headquarters'].keys() and len(str(string['headquarters']['city']))!=0:
##                                address = address + str(string['headquarters']['city'])+ ", "
##                            if 'state' in string['headquarters'].keys() and len(str(string['headquarters']['state']))!=0:
##                                address = address + str(string['headquarters']['state']) + " "
##                            if 'zip' in string['headquarters'].keys() and len(str(string['headquarters']['zip']))!=0:
##                                address = address + "-"+str(string['headquarters']['zip'])+ ", "
##                            if 'country' in string['headquarters'].keys():
##                                address = address + string['headquarters']['country']
##                            #address = str(string['headquarters'])
##                        else:
##                            address = "Address not available"
##                        if 'website' in string.keys():
##                            website = str(string['website'])
##                        else:
##                            website = "Website not available"
##                        if 'size' in string.keys():
##                            size = str(string['size'])
##                        else:
##                            size = "Company Size not available"
##                        if 'companyType' in string.keys():
##                            ctype = str(string['companyType'])
##                        else:
##                            ctype = "Company Type not available"
##                        if 'industry' in string.keys():
##                            industry = str(string['industry'])
##                        else:
##                            industry = "Industry not available"
##                        print(total_comp)
##                        print "Name: " + name
##                        print "Address: " + address
##                        print "Website: " + website
##                        print "Company Size: " + size
##                        print "Company Type: " + ctype
##                        print "Industry: " + industry
##                        print
##                        print
##                        print
##                        with open("games.txt", "a") as text_file:
##                             text_file.write("\n\n"+ "Name: " + name + "\n")
##                             text_file.write("Address: " + address + "\n")
##                             text_file.write("Website: " + website + "\n")
##                             text_file.write("Company Size: " + size+ "\n")
##                             text_file.write("Industry: " + industry + "\n\n")
##                        with open("games1.txt", "a") as text_file:
##                             text_file.write(name + "^"+address+"^"+website+"^"+size+"^"+industry+"\n")
##                             
                        
            

##                    book.save("compnames.xls")
                    
                          
                             

                    
                    
##                        for k,value in string['headquarters'].items():
##                            print(value)
##                            print(str(value))
##                            print(unicode(value))
##                            if type(value) == str:
##                                
##                                value = unicode(value, "utf-8", errors="ignore")
##                            else:
##
##                                value = unicode(value)
##                            #y = v.encode('ascii', 'ignore')
##                            print(k,"kkdk  ",value)
##                            #v = v.decode('unicode-escape')
##                            string['headquarters'][k] = value
                            
                        #string['headquarters'] = string['headquarters'].decode('unicode-escape')
                        #print(,": ",)
                        
                    #else:
                        #print(str(compname[0]), ": Headquarters not given")
                    
                        #except:
                            #pass
                        #print(string["yearFounded"])
                        
                        #print(s)
                        #ab=dict(e.split(":") for e in s.translate(None,"{}").split(","))
                        #print(ab)
                        #ab = eval(dict(s))
                        #print(ab)
                        #print(success_dict["yearFounded"])
                        #Now removing { and}
    ##                    s = ""
    ##                    s = string.replace("{" ,"");
    ##                    finalstring = s.replace("}" , "");
    ##
    ##                    #Splitting the string based on , we get key value pairs
    ##                    list1 = finalstring.split(",")
    ##                    #print(string)
    ##                    print()
    ##                    print(list1)
    ##
    ##                    dicta ={}
    ##                    for i in list1:
    ##                        #
    ##                        print(i)
    ##                        #Get Key Value pairs separately to store in dictionary
    ##                        keyvalue = i.split(":")
    ##                        print(keyvalue)
    ##                        #print(keyvalue[1])
    ##                        #Replacing the single quotes in the leading.
    ##                        m= keyvalue[0]
    ##                        #m = m.replace("\"", "")
    ##                        dicta[m] = keyvalue[1]
    ##
    ##                    print dicta
    ##                    ecount+=1
                
##                except HTTPError:
##                    print("HTTP ")
##                    #errcount+=1
##                    company_errors.append(url)
##                    
##                    continue
                    
                except UnicodeEncodeError:

                    pass

                except IndexError:
                    print("INDDDDDDDDDDEEEEEEEEEEEEEEEXXXXXXX",url)
           



##def crawl():
##    global alphanum1_errors
##    global alphanum2_errors
##    global alpha_errors
##    global current_letter
##    global current_num1
##    global current_num2
##    global randomvar
##    opener1 = ""
##    i = 0
##    j = 0
##    htmla =""
##    company_list= []
##    url_comp_list = []
##    linklist = []
##    count  = 1
##    xy = 1
##     
##    randomvar+=1
##    if randomvar == 1:
##        xy = 56
##    letter_list = ['b','c','d','e','f','g','h','i','j','k','l','m','n','o','p','q','r','s','t','u','v','w','x','y','z']
##    
##    for letter in letter_list:
##        current_letter = letter
##        print("Current letter: ",current_letter)
##        #try:
###            url1 = 'https://www.linkedin.com/directory/companies-' + letter+'/'
## #       url1 = 'https://www.linkedin.com/company-beta/19999997/'
## #       linklist.append(url1)
##        linklist = ['https://www.linkedin.com/company-beta/1035/','https://www.linkedin.com/company-beta/19999997/','https://www.linkedin.com/company-beta/19999996/','https://www.linkedin.com/company-beta/19999995/','https://www.linkedin.com/company-beta/19999994/','https://www.linkedin.com/company-beta/19999993/']
##        data(linklist)
##            
##            
####            resp1 = opener.open(url1)
####            print("thru")
####            cont1 = resp1.read()
####            doc1 = html.fromstring(cont1)
####            companies =  doc1.xpath('//li[@class="content"]/a/text()')
####            print(len(companies))
####            for count in range(xy,len(companies)):
####                try:
####                    print("First number : ",count)
####                    i = i+1
####                    url2 = 'https://www.linkedin.com/directory/companies-' + letter+'-'+str(count)+'/'
####                    #print(url2)
####                    resp2 = opener.open(url2)
####                    cont2 = resp2.read()
####                    doc2 = html.fromstring(cont2)
####                    companies1 =  doc2.xpath('//li[@class="content"]/a/text()')
####                    companies = companies + companies1
####                    if randomvar == 1:
####                        randomvar+=1
####                        count1 = 71
####                        while(count1<len(companies1)):
####                        #for count1 in range(28,):
####                            
####                            try:
####                                print("Second number: ", count1)
####                                print(alpha_errors)
####                                print(company_errors)
####                                print(alphanum1_errors)
####                                print(alphanum2_errors)
####                                j+=1
####                                url3 = 'https://www.linkedin.com/directory/companies-' + letter+'-'+str(count)+'-'+str(count1)+'/'
####                                print(url3)
####                                
####                                    
####                                resp3 = opener.open(url3)
####                                cont3 = resp3.read()
####                                doc3 = html.fromstring(cont3)
####                                companies2 =  doc3.xpath('//li[@class="content"]/a/@href')
####                                linklist=companies2
####                                data(linklist)
####                                count1+=1
####                                
####                            except HTTPError:
####                                print("prrrrrrrrrrrrobbbbbbblemm2222")
####                                alphanum2_errors.append(url3)
####                                continue
####                                
####                                
####                    
####                   
####                    else:
####                        count1 = 1
####                        while(count1<len(companies1)):
####                        #for count1 in range(1,len(companies1)):
####                            try:
####                                print("Second number: ", count1)
####                                print(alpha_errors)
####                                print(company_errors)
####                                print(alphanum1_errors)
####                                print(alphanum2_errors)
####                                j+=1
####                                url3 = 'https://www.linkedin.com/directory/companies-' + letter+'-'+str(count)+'-'+str(count1)+'/'
####                                print(url3)
####                                
####                                resp3 = opener.open(url3)
####                                cont3 = resp3.read()
####                                doc3 = html.fromstring(cont3)
####                                companies2 =  doc3.xpath('//li[@class="content"]/a/@href')
####                                linklist=companies2
####                                data(linklist)
####                                count1+=1
####                            except HTTPError:
####                                alphanum2_errors.append(url3)
####                                continue
####                except HTTPError:
####                    alphanum1_errors.append(url2)
####                    print("me1")
####                    pass
####                    #if j == 2:
####                       # break
####                #if i == 1:
####                    #break
####            #for k in link_list:
####                #print(k)
####            print("success")
####        except HTTPError:
####            alpha_errors.append(url1)
####            print("me")
####            continue
####    
##    








##        for company in company_list:
##            linklist.append('https://www.linkedin.com/company/'+str(url_comp_list[i]))
##            try:
##                alist = company.split()
##                alist = [x.lower() for x in alist]
##                if alist[-1][-1]!='.':
##                    for a in range(len(alist)):
##                        alist[a] = re.sub('[!@#$.,"()]', '', alist[a])
##                else:
##                    for a in range(len(alist)):
##                        alist[a] = re.sub('[!@#$.,"]()', '', alist[a])
##                    alist[-1] = str(alist[-1]) + '-'
##                url_comp_list.append('-'.join(alist))
##            except TypeError:
##                pass
##                
##        for i in range(len(url_comp_list)-1):
##            if url_comp_list[i] == url_comp_list[i+1]:
##                url_comp_list[i] = url_comp_list[i]+'_'+str(count)
##                count = count+1
##            else:
##                count = 1
##            
##            try:
##                linklist.append('https://www.linkedin.com/company/'+str(url_comp_list[i]))
##            except UnicodeEncodeError:
##                pass
        

            
##            
##            
            



            #url.addheaders = [{'User-Agent': "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.71 Safari/537.36"}]
            #url = "http://de.linkedin.com/directory/companies-a-3-1/"
            
            #htmla = requests.get(url).text
            #print htmla
            #print(j)
            #req.add_header({'User-Agent': "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.71 Safari/537.36"})
            #resp = urllib.urlopen(req)
            #content = resp.read()            
            #resp2.add_header('User-Agent', "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/39.0.2171.71 Safari/537.36")
            #resp2 = opener.open("https://www.linkedin.com/company/377948?trk=tyah&trkInfo=clickedVertical%3Acompany%2CentityType%3AentityHistoryName%2CclickedEntityId%3Acompany_377948%2Cidx%3A1")
                        
            
        
    #url = 'https://www.linkedin.com/company/3858'
##    resp2 = opener.open(
##    resp = opener.open('https://www.linkedin.com/directory/companies-computer-software/')
##    #printresp.read().xpath('//title'))
##    x = resp.read()
##    doc = html.fromstring(x)
##    alist = []
##    urllist = []
##    companylist = []
##    
##    
##    text1 =  doc.xpath('//li[@class="content"]/a/text()')
##    #print(text1)
##    for i in text1[0:-1]:
##        if i != u'S\xe3o Paulo - Computer Software/Engineering':
##            
##            #print(i)
##            y = i.split()
##            if len(y) == 4:
##                
##                y[0] = re.sub('[!@#$.,]', '', y[0])
##                alist.append(y[0].lower())
##            elif len(y) == 5:
##                y[0] = re.sub('[!@#$.,]', '', y[0])
##                y[1] = re.sub('[!@#$.,]', '', y[1])
##                alist.append(y[0].lower()+'-'+y[1].lower())
##            elif len(y) == 6:
##                y[0] = re.sub('[!@#$.,]', '', y[0])
##                y[1] = re.sub('[!@#$.,]', '', y[1])
##                y[2] = re.sub('[!@#$.,]', '', y[2])
##                alist.append(y[0].lower()+'-'+y[1].lower()+'-'+y[2].lower())
##            
##                
##            
##            
##    #print(alist)
##    for j in alist:
##        name = 'https://www.linkedin.com/directory/companies-'+j+'-computer-software/'
##        #print(name)
##        urllist.append(name)
##    #print(urllist)
##    for url in urllist:
##        resp1 = opener.open(url)
##        x1 = resp1.read()
##        doc1 = html.fromstring(x1)
##        text2 =  doc1.xpath('//li[@class="content"]/a/text()')
##        try:
##            for company in text2:
##                print(company)
##                companylist.append(company)
##                
##        except TypeError:
##            pass
##
##    #for i in companylist:
##       # print(i)
        
        
        
    
    #page = urllib.urlopen(url)
    #print(page)
    #print("okayyyyyyyyyyy")
    #x = ""
    #for l in page:
        
     #Filter it somehow
    #x = (resp.read())
    #print(x.xpath('//title'))
    # x = x + l
##                filename = 'yyyaj.html'
##                with open(filename, 'wb') as f:
##                    f.write(cont3)
    
    

def main():
    data()
    

if __name__ == '__main__':
    print("executing")
    main()





    
